import tkinter as tk
from tkinter import ttk
from tkinter import filedialog, messagebox
import pandas as pd
import os
import sys
import subprocess
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

RENAME_COLUMNS = {
   "остатки": {
        "Остаток на складе": "Остаток",
        "Себестоимость": "Себестоимость сумма",
        "Стоимость ( в розничных ценах)": "Сумма остатков в РЦ"
    },
    "продажи": {
        "Количество товаров": "Продажи",
        "Сумма продаж со скидкой": "Сумма продаж в РЦ"
    },
    "прайс": {
        "Номенклатура.Марка (Бренд)": "Бренд",
        "Номенклатура.Категория": "Категория",
        "Номенклатура.Сезон": "Сезон",
        "Цена (тг.)": "Прайс за ед.",
        "Марка (Бренд)": "Бренд"
    }
}

def find_header_row(path, keywords=["Магазин", "Номенклатура", "Характеристика"], max_rows=20):
    raw_df = pd.read_excel(path, header=None, nrows=max_rows)
    for i, row in raw_df.iterrows():
        values = row.astype(str).str.strip().tolist()
        if all(keyword in values for keyword in keywords):
            return i
    raise ValueError(f"Не удалось найти строку заголовков по ключевым словам: {keywords}")

def clean_file(path, тип_файла):
    try:
        header_row = find_header_row(path)
        if not isinstance(header_row, int):
            raise TypeError(f"Ожидался тип int для header_row, но получено: {type(header_row)}")
        print(f"Заголовки найдены на строке: {header_row}")
        df = pd.read_excel(path, header=header_row)
        print(f"Загружено: {df.shape[0]} строк, {df.shape[1]} столбцов")
        df.dropna(axis=0, how='all', inplace=True)
        df.dropna(axis=1, how='all', inplace=True)
        print(f"После очистки: {df.shape[0]} строк, {df.shape[1]} столбцов")
        rename_map = RENAME_COLUMNS.get(тип_файла, {})
        df = df.rename(columns=rename_map)
        print(f"Переименованные колонки: {list(df.columns)}")
        required = ["Магазин", "Номенклатура", "Характеристика"]
        missing = [col for col in required if col not in df.columns]
        if missing:
            raise ValueError(f"Отсутствуют обязательные колонки: {missing}")
        return df
    except Exception as e:
        print(f"Ошибка при обработке файла: {e}")
        raise

def format_excel_file(path: str, sheet_name: str = None, fixed_width: int = 10):
    wb = load_workbook(path)

    # выбираем лист по имени, если передан
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    if ws is None:
        raise ValueError("Файл не содержит активного листа")

    max_col = ws.max_column
    max_row = ws.max_row

    # 🔹 фиксированная высота заголовка
    ws.row_dimensions[1].height = 50

    # 🔹 автофильтр
    ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"

    # 🔹 закрепляем заголовок (строка 1)
    ws.freeze_panes = "A2"

    # 🔹 стиль границ
    thin = Side(border_style="thin", color="000000")

    for col in range(1, max_col + 1):
        col_letter = get_column_letter(col)

        # задаём фиксированную ширину (например, 20 символов)
        ws.column_dimensions[col_letter].width = fixed_width

        for row in range(1, max_row + 1):
            cell = ws[f"{col_letter}{row}"]

            if row == 1:  # заголовок
                cell.alignment = Alignment(vertical="center", wrap_text=True, horizontal="center")
            else:  # данные
                cell.alignment = Alignment(vertical="center", wrap_text=False)

            # рамка со всех сторон
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    wb.save(path)

class DataProcessor:
    def __init__(self):
        self.stock_df = None
        self.sales_df = None
        self.price_df = None

    def load_stock(self, path):
        self.stock_df = clean_file(path, тип_файла="остатки")

    def load_sales(self, path):
        self.sales_df = clean_file(path, тип_файла="продажи")

    def load_price(self, path):
        self.price_df = clean_file(path, тип_файла="прайс")

    def generate_summary(self):
        # Проверка загрузки всех таблиц
        if self.stock_df is None:
            raise ValueError("Файл с остатками не загружен")
        if self.sales_df is None:
            raise ValueError("Файл с продажами не загружен")
        if self.price_df is None:
            raise ValueError("Файл с прайсом не загружен")

        print("📦 Остатки:", self.stock_df.shape)
        print("📈 Продажи:", self.sales_df.shape)
        print("💰 Прайс:", self.price_df.shape)

        # Объединение таблиц
        df = self.stock_df.merge(
            self.sales_df,
            on=["Магазин", "Номенклатура", "Характеристика"],
            how="outer"
        )

        df_all = self.price_df.merge(
            df,
            on=["Магазин", "Номенклатура", "Характеристика"],
            how="left"
        )

        df_all = df_all.drop_duplicates(subset=["Магазин", "Номенклатура", "Характеристика"])

        columns_to_keep = [
            "Магазин", "Бренд", "Номенклатура", "Характеристика",
            "Категория", "Сезон",
            "Остаток", "Себестоимость сумма", "Продажи", "Прайс за ед.",
            "Сумма продаж в РЦ", "Сумма остатков в РЦ"
        ]

        missing = [col for col in columns_to_keep if col not in df_all.columns]
        if missing:
            raise ValueError(f"Отсутствуют ожидаемые колонки: {missing}")

        df_all = df_all[columns_to_keep]

        # Приводим к числовому типу
        df_all["Остаток"] = pd.to_numeric(df_all["Остаток"], errors="coerce")
        df_all["Себестоимость сумма"] = pd.to_numeric(df_all["Себестоимость сумма"], errors="coerce")

        # Расчет себестоимости за ед.
        df_all["Себестоимостьза ед."] = np.where(
            df_all["Остаток"] > 0,
            (df_all["Себестоимость сумма"] / df_all["Остаток"])
                .replace([np.inf, -np.inf], np.nan)
                .fillna(0)
                .round(0)
                .astype(int),
            0
        )

        desired_order = [
            "Магазин", "Бренд", "Категория", "Сезон",
            "Номенклатура", "Характеристика",
            "Остаток", "Себестоимость сумма", "Себестоимостьза ед.", "Продажи", "Прайс за ед.",
            "Сумма продаж в РЦ", "Сумма остатков в РЦ"
        ]

        df_all = df_all[desired_order]

        sort_columns = ["Магазин", "Бренд", "Категория", "Номенклатура", "Характеристика"]
        for col in sort_columns:
            if col not in df_all.columns:
                raise ValueError(f"Не хватает колонки для сортировки: {col}")

        df_all = df_all.sort_values(by=sort_columns)

        return df_all


from tkinter import ttk

class AppGUI:
    def __init__(self, root):
        self.root = root
        #self.root.title("Обработка остатков и продаж")
        self.root.geometry("480x420")
        self.processor = DataProcessor()

        # Заголовок
        title = ttk.Label(root, text="📊 Расспределение товара", font=("Arial", 14, "bold"))
        title.grid(row=0, column=0, columnspan=3, pady=10)

        # === Кнопки загрузки файлов ===
        ttk.Button(root, text="📦 Загрузить остатки", command=self.load_stock).grid(row=1, column=0, columnspan=3, sticky="ew", padx=20, pady=5)
        ttk.Button(root, text="📈 Загрузить продажи", command=self.load_sales).grid(row=2, column=0, columnspan=3, sticky="ew", padx=20, pady=5)
        ttk.Button(root, text="📋 Загрузить прайс-лист", command=self.load_price).grid(row=3, column=0, columnspan=3, sticky="ew", padx=20, pady=5)
        ttk.Button(root, text="⚙ Загрузить минимальные остатки", command=self.load_min_stock).grid(row=4, column=0, columnspan=3, sticky="ew", padx=20, pady=5)

        # === Поле ввода периода ===
        ttk.Label(root, text="Период прогноза (в днях):", font=("Arial", 11)).grid(row=5, column=0, sticky="e", padx=10, pady=10)
        self.days_entry = ttk.Entry(root, width=10)
        self.days_entry.insert(0, "14")
        self.days_entry.grid(row=5, column=1, sticky="w", padx=5, pady=10)

        # === Кнопки действий ===
        ttk.Button(root, text="📊 Рассчитать заказ и сохранить", command=self.calculate_order).grid(row=6, column=0, columnspan=3, sticky="ew", padx=20, pady=8)
        ttk.Button(root, text="📦 Подсорт с Центрального склада", command=self.save_distribution).grid(row=7, column=0, columnspan=3, sticky="ew", padx=20, pady=8)
        ttk.Button(root, text="🔄 Рассчитать межмаг", command=self.recalc_mezhmag).grid(row=8, column=0, columnspan=3, sticky="ew", padx=20, pady=8)
        ttk.Button(root, text="💾 Выгрузить межмаг", command=self.save_mezhmag_to_excel).grid(row=9, column=0, columnspan=3, sticky="ew", padx=20, pady=8)

        # Растягиваем центральную колонку
        root.grid_columnconfigure(0, weight=1)
        root.grid_columnconfigure(1, weight=1)
        root.grid_columnconfigure(2, weight=1)

    # === Загрузка файлов ===
    def load_stock(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if not path:
            return
        try:
            self.processor.load_stock(path) # здесь clean_file вызовется автоматически
            self.stock_df = self.processor.stock_df
            messagebox.showinfo("Файл загружен", f"Остатки загружены: {len(self.stock_df)} строк")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось загрузить остатки:\n{e}")

    def load_sales(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if not path:
            return
        try:
            self.processor.load_sales(path) # здесь clean_file вызовется автоматически
            self.sales_df = self.processor.sales_df
            messagebox.showinfo("Файл загружен", f"Продажи загружены: {len(self.sales_df)} строк")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось загрузить продажи:\n{e}")

    def load_price(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if not path:
            return
        try:
            self.processor.load_price(path) # здесь clean_file вызовется автоматически
            self.price_df = self.processor.price_df
            messagebox.showinfo("Файл загружен", f"Прайс-лист загружен: {len(self.price_df)} строк")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось загрузить прайс-лист:\n{e}")

    def load_min_stock(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if not path:
            return
        try:
            self.min_stock_df = pd.read_excel(path)
            if not {"Категория", "min stock", "max прием"}.issubset(self.min_stock_df.columns):
                raise ValueError("В файле должны быть колонки: Категория, min stock, max прием")
            messagebox.showinfo("Файл загружен", f"Минимальные остатки загружены: {len(self.min_stock_df)} строк")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось загрузить минимальные остатки:\n{e}")

    # === Расчёт заказа ===
    def calculate_order(self):
        if self.stock_df is None or self.sales_df is None or self.price_df is None:
            messagebox.showerror("Ошибка", "Сначала загрузите остатки, продажи и прайс-лист")
            return

        try:
            days = int(self.days_entry.get())
            if days <= 0:
                raise ValueError("Введите положительное число дней")
            
            # 🔹 передаём данные в процессор
            self.processor.stock_df = self.stock_df
            self.processor.sales_df = self.sales_df
            self.processor.price_df = self.price_df

            df = self.processor.generate_summary()  # Берём объединённый датафрейм через DataProcessor

            # Заменяем пустые значения на 0, чтобы не мешали при расчетах
            df["Остаток"] = df["Остаток"].fillna(0)
            df["Продажи"] = df["Продажи"].fillna(0)
            

            # Расчет заказа
            df["Заказ на период"] = df.apply(
                lambda row: (row.get("Продажи", 0) / 7.0) * days - row.get("Остаток", 0),
                axis=1
            )

            df["Заказ на период"] = df["Заказ на период"].fillna(0)

            # Комментарии
            def comment(row):
                if row["Остаток"] == 0 and row.get("Продажи", 0) == 0 and row["Заказ на период"] == 0:
                    return "Отправить минимальное количество"
                elif row["Заказ на период"] < 0:
                    return "Излишек"
                elif row["Заказ на период"] > 0:
                    return "Дозаказ"
                else:
                    return ""

            df["Комментарий"] = df.apply(comment, axis=1)

            # Минимальные остатки
            if self.min_stock_df is not None:
                df = df.merge(
                    self.min_stock_df[["Категория", "max прием"]],
                    on="Категория",
                    how="left"
                )
                df.loc[df["Комментарий"] == "Отправить минимальное количество", "Заказ на период"] = \
                    df.loc[df["Комментарий"] == "Отправить минимальное количество", "max прием"].fillna(0)
                df.drop(columns=["max прием"], inplace=True, errors="ignore")

            self.df_all = df

            # Сохраняем файл
            path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                title="Сохранить файл заказа"
            )
            if path:
                df.to_excel(path, index=False)
                messagebox.showinfo("Сохранено", f"Файл сохранен:\n{path}")
                try:
                    os.startfile(path)
                except Exception:
                    pass

        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось рассчитать заказ:\n{e}")

    # === Распределение с Центрального склада ===
    def safe_int(self, value):
        """Преобразует значение в целое число, заменяя NaN и ошибки на 0"""
        num = pd.to_numeric(value, errors="coerce")
        return int(0 if pd.isna(num) else num)


    def build_distribution(self, df):
        priority_stores = [
            "Гранд парк",
            "Азия парк Астана",
            "Шымкент «Love is mama»",
            "Aport East",
            "Aport West",
            "ГЦРЧ"
        ]
        
        # Создаем быстрый словарь для поиска строк
        lookup = df.set_index(["Магазин", "Номенклатура", "Характеристика"]).to_dict("index")
        

        central_df = df[df["Магазин"] == "Центральный склад"]
        result_rows = []

        for _, central_row in central_df.iterrows():
            central_stock = self.safe_int(central_row.get("Остаток", 0))

            row_data = {
                "Категория": central_row.get("Категория", ""),
                "Сезон": central_row.get("Сезон", ""),
                "Бренд": central_row.get("Бренд", ""),
                "Номенклатура": central_row.get("Номенклатура", ""),
                "Характеристика": central_row.get("Характеристика", ""),
                "Откуда": "Центральный склад",
                "Начальное кол-во у отправителя": central_stock,
            }

            for store in priority_stores:
                key = (store, central_row["Номенклатура"], central_row['Характеристика'])
                store_row = lookup.get(key)
                

                if store_row is None:
                    # row_data["{} Начальный остаток".format(store)] = 0
                    row_data["{} Количество заказа".format(store)] = 0
                    # row_data["{} Конечный остаток".format(store)] = 0
                else:
                    start_stock = self.safe_int(store_row.get("Остаток", 0))
                    comment = str(store_row.get("Комментарий", "")).strip()

                    if comment == "Дозаказ":
                        need = self.safe_int(store_row.get("Заказ на период", 0))
                    elif comment == "Отправить минимальное количество":
                        need = self.safe_int(store_row.get("Заказ на период", 0))
                    else:
                        need = 0


                    give = min(central_stock, need)
                    central_stock -= give
                    row_data["{} Количество заказа".format(store)] = give

            total_given = sum(row_data[f"{store} Количество заказа"] for store in priority_stores)

            # row_data["{} Начальный остаток".format(store)] = start_stock
            row_data["{} Количество заказа".format(store)] = give
            # row_data["{} Конечный остаток".format(store)] = start_stock + give
            row_data["Конечный остаток на ЦС"] = central_stock


            if total_given > 0:


                result_rows.append(row_data)
        result_df = pd.DataFrame(result_rows)

        #return pd.DataFrame(result_rows)
        # 🔹 Убираем строки, где на Центральном складе остаток = 0
        result_df = result_df[result_df["Начальное кол-во у отправителя"] > 0].reset_index(drop=True)

        return result_df
   

    def save_distribution(self):
        if self.df_all is None:
            messagebox.showerror("Ошибка", "Сначала рассчитайте заказ")
            return

        try:
            dist_df = self.build_distribution(self.df_all)
            self.distribution_df = dist_df
            path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                title="Сохранить файл Подсорт"
            )
            if not path:
                return

        # сохраняем Excel
            with pd.ExcelWriter(path, engine="openpyxl") as writer:
                 dist_df.to_excel(writer, index=False, sheet_name="Подсорт")

            # применяем форматирование
            format_excel_file(path)

            messagebox.showinfo("Сохранено", f"Файл сохранен:\n{path}")
            try:
                os.startfile(path)
            except Exception:
                pass

        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось построить распределение:\n{e}")

    def update_stock_after_distribution(self, df, dist_df):
    #Обновляем остатки магазинов по конечным остаткам из таблицы Подсорта
    # собираем обновления в единый DataFrame (быстрее, чем много итераций)
        updates = []
        for col in dist_df.columns:
            if "Конечный остаток" in col:
                # допускаем, что формат колонки: "<Store> Конечный остаток" (с пробелом)
                store_name = col.replace(" Конечный остаток", "").strip()
                temp = dist_df[["Номенклатура", "Характеристика", col]].copy()
                temp = temp.rename(columns={col: "Новый остаток"})
                temp["Магазин"] = store_name
                updates.append(temp)

        if not updates:
            return df

        updates_df = pd.concat(updates, ignore_index=True)

        # merge по ключам (фаст) и обновление остатков
        df = df.merge(updates_df, on=["Магазин", "Номенклатура", "Характеристика"], how="left")
        df["Остаток"] = df["Новый остаток"].fillna(df["Остаток"])
        df.drop(columns=["Новый остаток"], inplace=True, errors="ignore")

        return df


    
    def recalc_mezhmag(self):
        """
        Быстро формирует таблицу 'Межмаг' после подсорта:
        - обновляет остатки магазинов по данным из подсорта
        - пересчитывает заказ
        - обновляет комментарии
        - добавляет min stock / max приём
        - вызывает build_mezhmag_distribution
        """
        if self.df_all is None or self.distribution_df is None:
            messagebox.showerror("Ошибка", "Сначала рассчитайте Подсорт")
            return
        try:
            days = int(self.days_entry.get())
            if days <= 0:
                raise ValueError("Введите положительное число дней")

            df = self.df_all.copy()
            dist_df = self.distribution_df.copy()

            # === 1. Делаем «длинный» список конечных остатков по всем магазинам ===
            end_cols = [c for c in dist_df.columns if "Конечный остаток" in c]
            melted_frames = []
            for col in end_cols:
                store_name = col.replace(" Конечный остаток", "")
                tmp = dist_df.loc[:, ["Номенклатура", "Характеристика", col]].copy()
                tmp["Магазин"] = store_name
                tmp.rename(columns={col: "Конечный остаток"}, inplace=True)
                melted_frames.append(tmp)
            end_df = pd.concat(melted_frames, ignore_index=True)

            # === 2. Подставляем новые остатки магазинов через merge ===
            df = df.merge(
                end_df,
                on=["Магазин", "Номенклатура", "Характеристика"],
                how="left"
            )
            df["Остаток"] = df["Конечный остаток"].combine_first(df["Остаток"])
            df.drop(columns=["Конечный остаток"], inplace=True)

            # === 3. Пересчёт заказа (векторно) ===
            df["Продажи"] = df["Продажи"].fillna(0)
            df["Остаток"] = df["Остаток"].fillna(0)
            df["Заказ на период"] = (df["Продажи"] / 7.0) * days - df["Остаток"]

            # === 4. Комментарий (векторно) ===
            cond_min = (df["Остаток"] == 0) & (df["Продажи"] == 0) & (df["Заказ на период"] == 0)
            cond_excess = df["Заказ на период"] < 0
            cond_reorder = df["Заказ на период"] > 0
            df["Комментарий"] = np.select(
                [cond_min, cond_excess, cond_reorder],
                ["Отправить минимальное количество", "Излишек", "Дозаказ"],
                default=""
            )

            # === 5. Добавляем min stock / max прием (если есть) ===
            if self.min_stock_df is not None:
                df = df.merge(
                    self.min_stock_df[["Категория", "min stock", "max прием"]],
                    on="Категория",
                    how="left"
                )

            # === 6. Строим саму таблицу «Межмаг» по новой функции ===
            mezhmag_df = self.build_mezhmag_distribution(df)

            # === 7. Сохраняем результат в self ===
            self.mezhmag_df = mezhmag_df

            messagebox.showinfo("Готово", "Таблица Межмаг рассчитана.\nМожно выгрузить её кнопкой 'Выгрузить межмаг'")

        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось рассчитать межмаг:\n{e}")

        

    def save_mezhmag_to_excel(self):
        if not hasattr(self, "mezhmag_df") or self.mezhmag_df is None:
            messagebox.showerror("Ошибка", "Сначала рассчитайте межмаг")
            return

        try:
            path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                title="Сохранить файл Межмаг"
            )
            if not path:
                return

            # сохраняем Excel
            with pd.ExcelWriter(path, engine="openpyxl") as writer:
                self.mezhmag_df.to_excel(writer, index=False, sheet_name="Межмаг")

            # применяем форматирование
            try:
                format_excel_file(path)
            except Exception as fe:
                messagebox.showwarning("Внимание", f"Файл сохранен, но форматирование не применилось:\n{fe}")

            messagebox.showinfo("Сохранено", f"Файл сохранён:\n{path}")

            try:
                os.startfile(path)
            except Exception:
                pass

        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось сохранить файл:\n{e}")



    def build_mezhmag_distribution(self, df):
        priority_stores = [
            "Гранд парк",
            "Азия парк Астана",
            "Шымкент «Love is mama»",
            "Aport East",
            "Aport West",
            "ГЦРЧ"
        ]

        donor_priority = [
            "Aport East",
            "Aport West",
            "ГЦРЧ",
            "Гранд парк",
            "Шымкент «Love is mama»",
            "Азия парк Астана"
        ]

        # 🔹 Убираем Центральный склад
        df = df[df["Магазин"] != "Центральный склад"]

        # 🔹 Доноры: только из списка donor_priority
        donors = df[
            (df["Комментарий"] == "Излишек") &
            (df["Магазин"].isin(donor_priority))
        ].copy()

        donors["Доступно"] = (donors["Остаток"] - donors["min stock"].fillna(0)).clip(lower=0)
        donors["Доступно"] = np.minimum(donors["Доступно"], donors["Заказ на период"].abs())

        donors = donors[[
            "Магазин", "Категория", "Сезон", "Бренд",
            "Номенклатура", "Характеристика", "Остаток", "Доступно"
        ]]

        # 🔹 Получатели
        recipients = df[df["Комментарий"].isin(["Дозаказ", "Отправить минимальное количество"])].copy()
        recipients["Нужно"] = np.where(
            recipients["Комментарий"] == "Дозаказ",
            recipients["Заказ на период"],
            recipients["max прием"].fillna(0)
        )
        recipients = recipients[["Магазин", "Номенклатура", "Характеристика", "Нужно"]]

        # 🔹 Быстрые словари доступа
        df_lookup = df.set_index(["Магазин", "Номенклатура", "Характеристика"])["Остаток"].to_dict()
        rec_lookup = recipients.set_index(["Магазин", "Номенклатура", "Характеристика"])["Нужно"].to_dict()

        result_rows = []

        # 🔹 Проход по донорам в порядке приоритета
        donors = donors.sort_values(by="Магазин", key=lambda col: [donor_priority.index(m) for m in col])

        for _, donor in donors.iterrows():
            central_stock = donor["Доступно"]

            row_data = {
                "Категория": donor["Категория"],
                "Сезон": donor["Сезон"],
                "Бренд": donor["Бренд"],
                "Номенклатура": donor["Номенклатура"],
                "Характеристика": donor["Характеристика"],
                "Откуда": donor["Магазин"],
                "Начальное кол-во у отправителя": int(donor["Остаток"])
            }

            for store in priority_stores:
                key = (store, donor["Номенклатура"], donor["Характеристика"])
                start_stock = int(df_lookup.get(key, 0))
                rec_need = float(rec_lookup.get(key, 0))

                give = min(central_stock, rec_need)
                central_stock -= give

                # row_data[f"{store} Начальный остаток"] = start_stock
                row_data[f"{store} Количество заказа"] = give
                # row_data[f"{store} Конечный остаток"] = start_stock + give

            total_given = sum(row_data[f"{store} Количество заказа"] for store in priority_stores)
            if total_given > 0:
                row_data["Конечный остаток уотправителя"] = int(donor["Остаток"] - total_given)
                result_rows.append(row_data)

        return pd.DataFrame(result_rows)
    
    

if __name__ == "__main__":
    print("Старт программы")  # Проверка запуска
    root = tk.Tk()
    app = AppGUI(root)
    root.mainloop()
